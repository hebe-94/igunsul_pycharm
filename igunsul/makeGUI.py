from tkinter import *


def makeData():
    import requests
    import openpyxl
    from bs4 import BeautifulSoup as bs

    noticeList = []

    client = []
    announceNum = []
    announceName = []
    endDate = []
    price = []
    pricePercent = []

    Member_Data = {
        'status': 'login',
        'url2': 'common',
        'url': '',
        'login_page_value': '1',
        'go_page': '',
        'id': 'songam',
        'pw': 'qwer1234!@'
    }

    # 하나의 세션 객체를 생성해 일시적 유지
    with requests.Session() as s:
        requests = s.post('https://www.igunsul.net/login/login_process', data=Member_Data)

    # 나의방[엠에스그리드] 접근 URL
    requests = s.get(
        'http://igunsul.net/myroom?value="&freeze=&page_num=1&list_num=500&align=5&myroom_folder_eum=yes&folder_num=2&myroom_folder_info=&folder_seq=')

    print('myroom 내용')
    requests.encoding = 'euc-kr'
    soup = bs(requests.text, 'html.parser')
    result = soup.findAll('a', {"class": "list2detailAnchor listColormyroom_list myroom_list"})

    # 반복문 개수 결정(반복문 시작)
    print("공고 리스트 개수 : ", len(result))

    # 파싱으로 데이터 빼오기
    for a in range(0, len(result)):
        # 반복문 시작해서 다시 한번더 링크 변경
        requests = s.get('http://igunsul.net/' + result[a].get('href'))
        requests.encoding = 'euc-kr'
        soup = bs(requests.text, 'html.parser')

        noticeList.append('')
        client.append('')
        announceNum.append('')
        announceName.append('')
        endDate.append('')
        price.append('')
        pricePercent.append('')

        # client, announce 관련 정보 파싱
        clientResult = soup.find_all('td')

        client[a] = clientResult[2].text.replace('투찰금액산출', '').strip()
        announceNum[a] = clientResult[1].text.strip()
        announceName[a] = clientResult[0].text.strip()

        # print('clientResult값 확인')
        # x = 0
        # for rs in clientResult:
        #
        #     print(x, rs.text.strip())
        #     x = x + 1

        priceResult = soup.find('div', class_='rightNoticeSummaryContent')
        priceResult = priceResult.find_all('li')

        price[a] = priceResult[2].text.replace('원', '').strip()
        price[a] = price[a].replace('기초금액:', '').strip()
        pricePercent[a] = priceResult[4].text.strip()
        pricePercent[a] = pricePercent[a].replace('투찰률:', '').strip()
        endDate[a] = priceResult[5].text.strip()
        endDate[a] = endDate[a].replace('투찰마감일:', '').strip()

        # print('priceResult값 확인')
        #  가격정보 리스트 확인 (반복문)
        # x = 0
        # for rs in priceResult:
        #
        #     print(x, rs)
        #     x = x + 1

        print()
        print('////////////////////////산출 데이터///////////////////////')
        print('client : ', client[a])
        print('announceNum : ', announceNum[a])
        print('announceName : ', announceName[a])
        print('endDate : ', endDate[a])
        print('price : ', price[a])
        print('pricePercent : ', pricePercent[a])
        noticeList[a] = [client[a], announceNum[a], announceName[a], endDate[a], price[a], pricePercent[a]]
        print('notice리스트 : ', noticeList[a])
        print('///반복문 동작 횟수', a + 1)

    # 엑셀 load
    print('엑셀 데이터 표현')

    wb = openpyxl.load_workbook("C:/Users/SHIN/Desktop/python craw/python data.xlsx")
    ws = wb['Sheet1']

    # 데이터 넣기
    for row in range(4, len(result) + 4):
        for col in range(1, 7, 1):
            ws.cell(row=row, column=col).value = noticeList[row - 4][col - 1]

    # # 데이터 들어가는지 테스트
    # print(ws.cell(row=4, column=2).value)
    # print("%s" % ws.max_column)

    # 입력 데이터 console 확인
    # for row in range(4, len(result)+4):
    #     for col in range(1, 7, 1):
    #         print(ws.cell(row=row, column=col).value)
    #     print()

    wb.save('python data.xlsx')
    wb.close()

    # print('==============모든 행 열 출력===============')
    # all_values = []
    # for row in ws.rows:
    #     row_value = []
    #     for cell in row:
    #         row_value.append(cell.value)
    #     all_values.append(row_value)
    #
    # for a in all_values:
    #     print(a)

    #     writeWs['A'+a.__str__()] = noticeList[index]

    # loadWs["A"+a.__str__()].value = 'a'
    # print(loadWs["A"+a.__str__()].value)


root = Tk()
root.title("세훈이의 공고정리")
root.geometry("400x400")
root.resizable(False, False)

# 엑셀 데이터 만들기
dataToExcelBtn = Button(root, width=40, height=2, text="엠에스그리드 공고리스트 Excel 파일 만들기", command=makeData)
dataToExcelBtn.pack()

# 인콘 데이터 확인
dataNoticeBtn = Button(root, width=40, height=2, text="인콘 데이터 알림")
dataNoticeBtn.pack()

root.mainloop()
